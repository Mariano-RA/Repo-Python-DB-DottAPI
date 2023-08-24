from flask import Flask, request, jsonify
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.utils import FORMULAE
import requests
import xlwings as xw
import pandas as pd
import csv
import json
import os

app = Flask(__name__)

# Direccion archivos
listadosTemporales = "archivosPorCargar\\temporales\\"
listadoCsv = "archivosPorCargar\\csv"
listadoJson = "archivosPorCargar\\archivosJson\\"
diccionarios = "nuevosScripts\diccionarios\diccionarios.json"

# Función para encontrar un valor en el diccionario


def encontrar_valor(diccionario, clave):
    if clave in diccionario:
        return diccionario[clave]
    else:
        return "No existe una categoría para este producto"

# Función para obtener un diccionario


def obtenerDiccionario(nombreDiccionario):
    with open(diccionarios) as diccionariosOpen:
        diccionariosJson = json.load(diccionariosOpen)
    diccionarioBuscado = diccionariosJson[nombreDiccionario]
    return diccionarioBuscado


# TOKEN
access_token = "eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCIsImtpZCI6ImN5VnNWTzY4Z2hJNFZEUC1aYkpxMyJ9.eyJodHRwOi8vbG9jYWxob3N0OjMwMDAvcm9sZXMiOlsiYWRtaW4iXSwiaXNzIjoiaHR0cHM6Ly9kZXYtM2hteTNnYWxtcWZ0eTF4dC51cy5hdXRoMC5jb20vIiwic3ViIjoiZmFjZWJvb2t8MzMzNzI3MjU5MDc1ODMzIiwiYXVkIjpbImh0dHBzOi8vZG90dC1wYy1zZXJ2ZXIuY29tIiwiaHR0cHM6Ly9kZXYtM2hteTNnYWxtcWZ0eTF4dC51cy5hdXRoMC5jb20vdXNlcmluZm8iXSwiaWF0IjoxNjkyODkxMjMyLCJleHAiOjE2OTI5Nzc2MzIsImF6cCI6IjhON3NpcWR2cDVXbHhJVU5DM2hXS1IxUWlXTTkyMjIzIiwic2NvcGUiOiJvcGVuaWQgcHJvZmlsZSBlbWFpbCIsInBlcm1pc3Npb25zIjpbImNyZWF0ZTp0YWJsYXMiLCJyZWFkOmRhdG9zIl19.Ky-16IjDMhyoIcy1Ueg7za1TQs1zI8EQlqbD5bQUEDEUMK7gWhfkNRVEF3Uy_p4SIb4hSd3SWrnw0ptxZG1gpXVQDlWV_g2VbENtWhijwNvbRH27B2R3IRvN0u6NseiZROL2zWD6UVZo8UCwADTHtqNoBbiP0iE8_Py64PQG-AsUVAMcSksEIgd5GKcTZ_Nvht4s1oCMOmZINpcK1lgIBWCPAS9neGBOxfOjyMezyEDMjq5pKaaF2mBokJMnMps0j80MiToQjFhQWlRbTEww6rbHIZ8BKn6w-Qcq6QXjNEjzLy8K8lC1Cz6-oIDRJsWAaj5fLp4VqbJoWMheffc1_g"
url = "http://localhost:3000/api/productos/"
headers = {
    "Authorization": f"Bearer {access_token}"
}


@app.route("/procesar_archivo_air", methods=["POST"])
def procesar_archivo_air():

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    # Guardar el archivo temporalmente
    archivo.save(listadosTemporales+"air.csv")

    # Función para crear el archivo JSON

    with open(listadosTemporales+"air.csv", "r") as file:
        csv_reader = csv.reader(file, delimiter=",")
        next(csv_reader)  # Ignora la primera fila de encabezados
        data = []
        for row in csv_reader:
            if row[2] != "A":
                descripcion = row[1]
                rubro = row[10]
                iva = row[4]
                precio = row[2]
                registro = {
                    "proveedor": "air",
                    "producto": descripcion,
                    "categoria": encontrar_valor(obtenerDiccionario("air"), rubro),
                    "precio": round((float(precio) * (1 + (float(iva) / 100)) * 1.1))
                }
                data.append(registro)

    response = requests.post(url=url, json=data, headers=headers)

    if response.status_code == 201:
        return jsonify({"message": "Se actualizo la tabla de air correctamente"})
    else:
        return jsonify({"message": "No se pudo actualizar la tabla"})


@app.route("/procesar_archivo_eikon", methods=["POST"])
def procesar_archivo_eikon():

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    archivo.save(listadosTemporales+"temp_eikon.xlsx")

    df = pd.read_excel(listadosTemporales+"temp_eikon.xlsx")

    # Se borran las primeras filas
    df = df.drop([0, 1, 2])
    df.reset_index(drop=True, inplace=True)

    # Guardar como CSV
    df.to_csv(listadoCsv+"listadoEik.csv", index=False)

    with open(listadoCsv+"listadoEik.csv", "r") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:

            descripcion = row[1]
            categoria = row[6]
            precio = row[3]

            # Crea un diccionario con los datos de cada registro
            registro = {
                "proveedor": "eikon",
                "producto": descripcion,
                "categoria": encontrar_valor(obtenerDiccionario("eikon"), categoria),
                "precio": round((float(precio) * 1.1))
            }

            # Agrega el diccionario a la lista de datos
            data.append(registro)

    response = requests.post(url=url, json=data, headers=headers)

    if response.status_code == 201:
        return jsonify({"message": "Se actualizo la eikon de eikon correctamente"})
    else:
        return jsonify({"message": "No se pudo actualizar la tabla"})


@app.route("/procesar_archivo_elit", methods=["POST"])
def procesar_archivo_elit():

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    archivo.save(listadosTemporales+"temp_elit.xlsx")
    df = pd.read_excel(listadosTemporales+"temp_elit.xlsx")
    df.to_csv(listadoCsv+"listadoElit.csv", index=False)

    # Abre el archivo CSV en modo lectura con la codificación adecuada
    with open(listadoCsv+"listadoElit.csv", "r") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:
            descripcion = row[1]
            categoria = row[5]
            precio = row[7]
            iva = row[8]
            ivaInterno = row[9]

            # Crea un diccionario con los datos de cada registro
            registro = {
                "proveedor": "elit",
                "producto": descripcion,
                "categoria": encontrar_valor(obtenerDiccionario("elit"), categoria),
                "precio": round((float(precio) * (1 + (float(iva) + float(ivaInterno))/100) * 1.1))
            }

            # Agrega el diccionario a la lista de datos
            data.append(registro)

    response = requests.post(url=url, json=data, headers=headers)

    if response.status_code == 201:
        return jsonify({"message": "Se actualizo la tabla de elit correctamente"})
    else:
        return jsonify({"message": "No se pudo actualizar la tabla"})


def obtenerTipoIva(clave):
    # Diccionario con tipo de IVA
    tipoIva = {
        "002-I.V.A. 10.5 %": 10.5,
        "001-I.V.A. 21 %": 21,
        "005-Impuestos Internos": 21
    }
    if clave in tipoIva:
        return tipoIva[clave]


@app.route("/procesar_archivo_hdc", methods=["POST"])
def procesar_archivo_hdc():

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    archivo.save(listadosTemporales+"temp_hdc.xlsx")
    df = pd.read_excel(listadosTemporales+"temp_hdc.xlsx")
    df.to_csv(listadoCsv+"listadoHdc.csv", index=False)

    # Abre el archivo CSV en modo lectura con la codificación adecuada
    with open(listadoCsv+"listadoHdc.csv", "r", encoding="utf-8") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        next(csv_reader)  # Ignora la primera fila de encabezados

        for row in csv_reader:
            if row[7]:
                descripcion = row[5]
                if row[3]:
                    categoria = row[3]
                else:
                    categoria = row[2]
                precio = row[7]
                iva = obtenerTipoIva(row[8])

                # Crea un diccionario con los datos de cada registro
                registro = {
                    "proveedor": "hdc",
                    "producto": descripcion,
                    "categoria": encontrar_valor(obtenerDiccionario("hdc"), categoria),
                    "precio": round(float(precio) * (1+float(iva)/100) * 1.1)
                }

                # Agrega el diccionario a la lista de datos
                data.append(registro)

    response = requests.post(url=url, json=data, headers=headers)

    if response.status_code == 201:
        return jsonify({"message": "Se actualizo la tabla de hdc correctamente"})
    else:
        return jsonify({"message": "No se pudo actualizar la tabla"})


@app.route("/procesar_archivo_invid", methods=["POST"])
def procesar_archivo_invid():

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    # Guardar el archivo .xls en el directorio
    # archivo.save(listadosTemporales + "temp_invid.xlsx")
    df = pd.read_excel(archivo)

    # Se utilizan los índices 0, 1 y 2 para las primeras tres filas
    df = df.drop([0, 1, 2, 3, 4, 5, 6, 7, 8])
    df.reset_index(drop=True, inplace=True)
    df.to_excel(listadosTemporales + "temp_invid.xlsx", index=False)

    # Cargar el libro de trabajo
    book = load_workbook(listadosTemporales + "temp_invid.xlsx")
    sheet = book["Sheet1"]

    sheet["I3"] = "categoria"
    # Obtén la letra de la columna "I"
    column_letter = get_column_letter(9)

    # Aplica la fórmula en las celdas de la columna "I" (desde la fila 5 en adelante)
    for row_num in range(4, sheet.max_row + 1):
        formula_row = f'=IF(LEN(H{row_num})<2,"",IF(AND(LEN(A{row_num-2})<2,LEN(C{row_num-2})<2),B{row_num-2},{column_letter}{row_num-1}))'
        cell_reference = f"{column_letter}{row_num}"
        sheet[cell_reference] = formula_row

    book.save(listadosTemporales + "temp_invid.xlsx")
    book.close()

    app = xw.App(visible=False)  # Abre Excel en segundo plano
    libro = app.books.open(listadosTemporales + "temp_invid.xlsx")
    hoja = libro.sheets["Sheet1"]

    # Encuentra el último número de fila en la columna "I"
    ultima_fila = hoja.range("I" + str(hoja.cells.last_cell.row)).end("up").row

    # Empieza en la fila 2 para omitir encabezados
    for fila_numero in range(2, ultima_fila + 1):
        celda_formula = hoja.range("I" + str(fila_numero))
        resultado = celda_formula.value
        celda_formula.value = resultado

    # Cierra el archivo Excel
    libro.save()
    libro.close()
    app.quit()

    # Guardar como CSV
    df = pd.read_excel(listadosTemporales+"temp_invid.xlsx")
    df.to_csv(listadoCsv+"listadoInvid.csv", index=False)

    with open(listadoCsv+"listadoInvid.csv", "r") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        # Lee cada fila del archivo CSV (ignorando la primera fila de encabezados)

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:
            if (row[3] != "" and row[3] != "Nro. de Parte"):
                descripcion = row[1]
                categoria = row[8]
                precio = float(row[5])
                iva = (1 + (float(row[6])/100)) * (1 + (float(row[7])/100))

                # Crea un diccionario con los datos de cada registro
                registro = {
                    "proveedor": "invid",
                    "producto": descripcion,
                    "categoria": encontrar_valor(obtenerDiccionario("invid"), categoria),
                    "precio": round((precio * iva * 1.1))
                }

                # Agrega el diccionario a la lista de datos

                data.append(registro)

    response = requests.post(url=url, json=data, headers=headers)

    if response.status_code == 201:
        return jsonify({"message": "Se actualizo la tabla de invid correctamente"})
    else:
        return jsonify({"message": "No se pudo actualizar la tabla"})


@app.route("/procesar_archivo_nb", methods=["POST"])
def procesar_archivo_nb():

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    # Guardar el archivo temporalmente
    archivo.save(listadoCsv+"listadoNb.csv")

    # Abre el archivo CSV en modo lectura con la codificación adecuada
    with open(listadoCsv+"listadoNb.csv", "r", encoding="utf8") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=";")

        # Crea una lista para almacenar los datos
        data = []

        # Lee cada fila del archivo CSV (ignorando la primera fila de encabezados)

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:
            descripcion = row[3]
            categoria = row[2]
            precio = row[10]

            # Crea un diccionario con los datos de cada registro
            registro = {
                "proveedor": "nb",
                "producto": descripcion,
                "categoria": encontrar_valor(obtenerDiccionario("nb"), categoria),
                "precio": round((float(precio) * 1.1))
            }

            # Agrega el diccionario a la lista de datos
            data.append(registro)

    response = requests.post(url=url, json=data, headers=headers)

    if response.status_code == 201:
        return jsonify({"message": "Se actualizo la tabla de nb correctamente"})
    else:
        return jsonify({"message": "No se pudo actualizar la tabla"})


@app.route("/procesar_archivo_mega", methods=["POST"])
def procesar_archivo_mega():

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    # Guardar el archivo .xls en el directorio
    # archivo.save(listadosTemporales + "temp_megaTest.xlsx")
    df = pd.read_excel(archivo)

    # Se utilizan los índices 0, 1 y 2 para las primeras tres filas
    df = df.drop([0, 1])
    df.reset_index(drop=True, inplace=True)
    df.to_excel(listadosTemporales + "tempMega.xlsx", index=False)

    # Cargar el libro de trabajo
    book = load_workbook(listadosTemporales + "tempMega.xlsx")
    sheet = book["Sheet1"]

    # Define the base formula
    formula_base = '=IF(AND(ISBLANK(B{}), ISBLANK(B{})), CONCATENATE(A{}, " - \
        ", A{}), IF(ISBLANK(E{}), "", IF(LEN(B{}) < 2, CONCATENATE(LEFT(G{}, FIND(" - \
        ", G{}) - 1), " - ", A{}), G{})))'

    # Apply the formula to cells in column G, starting from G4
    for row in range(4, sheet.max_row + 1):
        row_references = row  # Adjust row references in the formula
        current_formula = formula_base.format(row_references-2, row_references-1, row_references-2, row_references-1,
                                              row_references, row_references-1, row_references-2, row_references-2, row_references-1, row_references-1)

        cell = sheet.cell(row=row, column=7)  # Column G
        cell.value = current_formula

    # Save the changes to the file
    book.save(listadosTemporales + "tempMega.xlsx")

    app = xw.App(visible=False)  # Abre Excel en segundo plano
    libro = app.books.open(listadosTemporales + "tempMega.xlsx")
    hoja = libro.sheets["Sheet1"]

    # Encuentra el último número de fila en la columna "I"
    ultima_fila = hoja.range("G" + str(hoja.cells.last_cell.row)).end("up").row

    # Empieza en la fila 2 para omitir encabezados
    for fila_numero in range(4, ultima_fila + 1):
        celda_formula = hoja.range("G" + str(fila_numero))
        resultado = celda_formula.value
        celda_formula.value = resultado

    # # Cierra el archivo Excel
    libro.save()
    libro.close()
    app.quit()

    # Guardar como CSV
    df = pd.read_excel(listadosTemporales+"tempMega.xlsx")
    df.to_csv(listadoCsv+"listadoMega.csv", index=False)

    # Abre el archivo CSV en modo lectura con la codificación adecuada
    with open(listadoCsv+"listadoMega.csv", "r", encoding="utf-8") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        # Lee cada fila del archivo CSV (ignorando la primera fila de encabezados)

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:
            if (row[3] != ""):
                descripcion = row[1]
                precio = row[2].replace("U$s ", "")
                iva = row[4].replace("+", "").replace("%", "")
                categoria = row[6]

                # Crea un diccionario con los datos de cada registro
                registro = {
                    "proveedor": "mega",
                    "producto": descripcion,
                    "categoria": encontrar_valor(obtenerDiccionario("mega"), categoria),
                    "precio": round((float(precio) * (1 + (float(iva)/100)) * 1.1))
                }

                # Agrega el diccionario a la lista de datos

                data.append(registro)

    # with open(listadoJson+"listadoMega.json", "w") as jf:
    #     json.dump(data, jf, ensure_ascii=False, indent=2)

    response = requests.post(url=url, json=data, headers=headers)

    if response.status_code == 201:
        return jsonify({"message": "Archivo mega procesado correctamente"})
    else:
        return jsonify({"message": "No se pudo actualizar la tabla"})


if __name__ == "__main__":
    app.run(debug=True)
